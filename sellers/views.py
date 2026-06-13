import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from products.models import Product
from orders.models import OrderItem
from .forms import SellerForm, SellerProductForm
from .models import Seller

MONTH_NAMES_RU = [
    '', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
    'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь',
]


def _get_seller(user):
    return Seller.objects.filter(user=user).first()


def _paid_items(seller):
    return OrderItem.objects.filter(product__seller=seller, order__status='paid')


@login_required
def become_seller(request):
    """Регистрация пользователя как продавца (создание магазина)."""
    seller = _get_seller(request.user)
    if seller:
        return redirect('seller_dashboard')

    if request.method == 'POST':
        form = SellerForm(request.POST)
        if form.is_valid():
            seller = form.save(commit=False)
            seller.user = request.user
            seller.save()
            messages.success(request, 'Магазин создан. Теперь вы можете добавлять товары.')
            return redirect('seller_dashboard')
    else:
        form = SellerForm()
    return render(request, 'sellers/become_seller.html', {'form': form})


@login_required
def seller_dashboard(request):
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')

    products = seller.products.select_related('category').order_by('-created_at')

    # Выручка продавца — по оплаченным заказам
    paid_items = OrderItem.objects.filter(
        product__seller=seller, order__status='paid'
    )
    stats = paid_items.aggregate(
        revenue=Sum(F('price') * F('quantity'), output_field=DecimalField()),
        sales_count=Sum('quantity'),
    )
    return render(request, 'sellers/dashboard.html', {
        'seller': seller,
        'products': products,
        'products_count': products.count(),
        'revenue': stats['revenue'] or 0,
        'sales_count': stats['sales_count'] or 0,
    })


@login_required
def edit_shop(request):
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')
    if request.method == 'POST':
        form = SellerForm(request.POST, instance=seller)
        if form.is_valid():
            form.save()
            messages.success(request, 'Данные магазина обновлены.')
            return redirect('seller_dashboard')
    else:
        form = SellerForm(instance=seller)
    return render(request, 'sellers/edit_shop.html', {'form': form, 'seller': seller})


@login_required
def product_create(request):
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')
    if request.method == 'POST':
        form = SellerProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.seller = seller
            product.save()
            messages.success(request, 'Товар добавлен.')
            return redirect('seller_dashboard')
    else:
        form = SellerProductForm()
    return render(request, 'sellers/product_form.html', {'form': form, 'title': 'Новый товар'})


@login_required
def product_edit(request, pk):
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')
    product = get_object_or_404(Product, pk=pk, seller=seller)
    if request.method == 'POST':
        form = SellerProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Товар обновлён.')
            return redirect('seller_dashboard')
    else:
        form = SellerProductForm(instance=product)
    return render(request, 'sellers/product_form.html', {
        'form': form, 'title': 'Редактирование товара', 'product': product,
    })


@login_required
def product_delete(request, pk):
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')
    product = get_object_or_404(Product, pk=pk, seller=seller)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Товар удалён.')
        return redirect('seller_dashboard')
    return render(request, 'sellers/product_confirm_delete.html', {'product': product})


@login_required
def seller_sales(request):
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')

    paid_items = _paid_items(seller)
    items = paid_items.select_related(
        'product', 'order', 'order__user'
    ).order_by('-order__created_at')
    total = paid_items.aggregate(
        revenue=Sum(F('price') * F('quantity'), output_field=DecimalField())
    )['revenue'] or 0

    # Помесячная выручка для графика
    monthly = (
        paid_items
        .annotate(month=TruncMonth('order__created_at'))
        .values('month')
        .annotate(revenue=Sum(F('price') * F('quantity'), output_field=DecimalField()))
        .order_by('month')
    )
    chart_labels = [m['month'].strftime('%m.%Y') for m in monthly]
    chart_values = [float(m['revenue']) for m in monthly]

    # Месяцы, за которые есть продажи — для выгрузки отчёта
    report_months = [
        {'value': m['month'].strftime('%Y-%m'),
         'label': f"{MONTH_NAMES_RU[m['month'].month]} {m['month'].year}"}
        for m in reversed(list(monthly))
    ]

    return render(request, 'sellers/sales.html', {
        'seller': seller,
        'items': items,
        'total': total,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
        'report_months': report_months,
        'current_month': timezone.now().strftime('%Y-%m'),
    })


@login_required
def seller_report(request):
    """Выгрузка отчёта о продажах за выбранный месяц в формате Excel (.xlsx)."""
    seller = _get_seller(request.user)
    if not seller:
        return redirect('become_seller')

    month_param = request.GET.get('month', '')
    try:
        period = datetime.strptime(month_param, '%Y-%m')
    except ValueError:
        period = timezone.now()

    items = _paid_items(seller).filter(
        order__created_at__year=period.year,
        order__created_at__month=period.month,
    ).select_related('product', 'order', 'order__user').order_by('order__created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Продажи'

    title = f'Отчёт о продажах — {MONTH_NAMES_RU[period.month]} {period.year}'
    ws.append([title])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    headers = ['Дата', 'Заказ', 'Товар', 'Покупатель', 'Кол-во', 'Цена, ₽', 'Сумма, ₽']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    total = 0
    for item in items:
        line_total = item.price * item.quantity
        total += line_total
        ws.append([
            item.order.created_at.strftime('%d.%m.%Y'),
            f'№{item.order.id}',
            item.product.title,
            item.order.user.username,
            item.quantity,
            float(item.price),
            float(line_total),
        ])

    ws.append([])
    ws.append(['', '', '', '', '', 'Итого:', float(total)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    widths = [12, 10, 40, 18, 10, 12, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'sales_{period.strftime("%Y_%m")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def seller_shop(request, slug):
    """Публичная витрина магазина продавца."""
    seller = get_object_or_404(Seller, slug=slug, is_active=True)
    products = seller.products.filter(is_active=True).select_related('category')
    return render(request, 'sellers/shop.html', {'seller': seller, 'products': products})
